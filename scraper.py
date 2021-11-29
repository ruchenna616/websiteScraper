#!/usr/bin/env python 3

#for reading excel files
import openpyxl
from pathlib import Path

#for parsing websites
from bs4 import BeautifulSoup
from urllib.request import urlopen
import re

#reading excel file
file = input("File Name: ")
word = input("Word To Find: ")

def readExcelFile(file):
    xlsx_file = Path(file)
    #opens workbook in excel file
    workbook = openpyxl.load_workbook(xlsx_file)
    #makes sheet available to use
    sheet = workbook.active

    #read each first element of each row
    row_names = []
    for row in sheet.iter_rows(1, sheet.max_row):
        row_names.append(row[0].value)

    return row_names

def parseWebsite(url):
    #gets html file for wbesite
    html = urlopen(url)
    #make "soup" html from https 
    soup = BeautifulSoup(html.read(), 'html.parser')

    #all text in website page
    text = []
    paragraphs = soup.find_all('p')
    #h1 - h6 headers
    headers = soup.find_all(re.compile('^h[1-6]$'))

    for para in paragraphs:
        text.append(str(para))
    for head in headers:
        text.append(str(head))

    #finds number of times the word is found in text array
    found = 0 
    for line in text:
        lineStr = str(line)
        index = lineStr.find(word)
        if(index != -1):
            found += 1

    return found
#main
websites = readExcelFile(file)
#parses each website in excel file
for site in websites:
    times_found = parseWebsite(site)
    print("WEBSITE: " + str(site) + "\n The word [" + word + "] was found " + str(times_found) + " times")
